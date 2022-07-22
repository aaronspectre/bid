from main.models import Bid, Channel
from openpyxl import Workbook
from django.urls import reverse
from django.shortcuts import render
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib.auth import authenticate, login, logout
from django.conf import settings


def index(request):
	if request.user.is_authenticated:
		return HttpResponseRedirect(reverse('dashboard'))
	return render(request, 'index.html')


def sign_in(request):
	user = authenticate(request, username = request.POST['username'], password = request.POST['password'])
	if user is not None:
		login(request, user)
		return HttpResponseRedirect(reverse('dashboard'))
	else:
		return HttpResponseRedirect(reverse('index'))


def sign_out(request):
	logout(request)
	return HttpResponseRedirect(reverse('index'))


def dashboard(request):
	if request.user.is_anonymous:
		return HttpResponseRedirect(reverse('index'))

	bids = Bid.objects.filter(status = 'alive').order_by('-date')
	return render(request, 'dashboard.html', {'bids': bids})


def download(request):
	bids = Bid.objects.all()

	workbook = Workbook()
	worksheet = workbook.active
	worksheet.title = 'Applications'

	for index in range(1, bids.count()+1):
		worksheet.cell(row = index, column = 1, value = bids[index-1].name)
		worksheet.cell(row = index, column = 2, value = bids[index-1].phone)
		worksheet.cell(row = index, column = 3, value = bids[index-1].language)
		worksheet.cell(row = index, column = 4, value = bids[index-1].source)
		worksheet.cell(row = index, column = 5, value = bids[index-1].product)
		workbook.save(f'{settings.BASE_DIR}/applications.xlsx')

	with open(f'{settings.BASE_DIR}/applications.xlsx', 'rb') as file:
		response = HttpResponse(file.read(), content_type = 'application/vnd.ms-excel')
		response['Content-Disposition'] = 'inline; filename=applications.xlsx'
		return response



def history(request):
	if request.user.is_anonymous:
		return HttpResponseRedirect(reverse('index'))

	bids = Bid.objects.filter(status = 'history').order_by('-date')
	return render(request, 'dashboard.html', {'bids': bids})



def remove(request, id):
	bid = Bid.objects.get(pk = id)
	bid.status = 'history'
	bid.save()
	return HttpResponseRedirect(reverse('dashboard'))



def channels(request):
	channels = Channel.objects.all()
	return render(request, 'channels.html', {'channels': channels})


def channels_add(request):
	channel = Channel()
	channel.name = request.POST['name']
	channel.save()
	return HttpResponseRedirect(reverse('channels'))


def channels_remove(request, id):
	Channel.objects.get(pk = id).delete()
	return HttpResponseRedirect(reverse('channels'))